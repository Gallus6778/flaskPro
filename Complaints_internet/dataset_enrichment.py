import xml.etree.cElementTree as ET
import openpyxl

def put_data_in_dataset(filename, msisdn_info_results):
    
    """ Return dictionnary of xml value"""

    imsi = None
    encKey = None
    algoId = None
    kdbId = None
    acsub = None
    imsiActive = None
    accTypeGSM = None
    accTypeGERAN = None
    accTypeUTRAN = None
    odboc = None
    odbic = None
    odbr = None
    odboprc = None
    odbssm = None
    odbgprs = None
    odbsci = None
    isActiveIMSI = None
    msisdn = None
    actIMSIGprs = None
    obGprs = None
    qosProfile = None
    refPdpContextName = None
    imeisv = None
    ldapResponse = None

    # Extraction du contenu du fichier xml
    tree = ET.ElementTree(file=filename)
    root = tree.getroot()
    inc = 1
    for books in root.findall('.//'):

        if (books.tag == 'imsi'):
            msisdn_info_results[books.tag] = books.text
            imsi = books.text

        if (books.tag == 'encKey'):
            msisdn_info_results[books.tag] = books.text
            encKey = books.text

        if (books.tag == 'algoId'):
            msisdn_info_results[books.tag] = books.text
            algoId = books.text

        if (books.tag == 'kdbId'):
            msisdn_info_results[books.tag] = books.text
            kdbId = books.text

        if (books.tag == 'acsub'):
            msisdn_info_results[books.tag] = books.text
            acsub = books.text

        if (books.tag == 'imsiActive'):
            msisdn_info_results[books.tag] = books.text
            imsiActive = books.text

        if (books.tag == 'accTypeGSM'):
            msisdn_info_results[books.tag] = books.text
            accTypeGSM = books.text

        if (books.tag == 'accTypeGERAN'):
            msisdn_info_results[books.tag] = books.text
            accTypeGERAN = books.text

        if (books.tag == 'accTypeUTRAN'):
            msisdn_info_results[books.tag] = books.text
            accTypeUTRAN = books.text

        if (books.tag == 'odboc'):
            msisdn_info_results[books.tag] = books.text
            odboc = books.text

        if (books.tag == 'odbic'):
            msisdn_info_results[books.tag] = books.text
            odbic = books.text

        if (books.tag == 'odbr'):
            msisdn_info_results[books.tag] = books.text
            odbr = books.text

        if (books.tag == 'odboprc'):
            msisdn_info_results[books.tag] = books.text
            odboprc = books.text

        if (books.tag == 'odbssm'):
            msisdn_info_results[books.tag] = books.text
            odbssm = books.text

        if (books.tag == 'odbgprs'):
            msisdn_info_results[books.tag] = books.text
            odbgprs = books.text

        if (books.tag == 'odbsci'):
            msisdn_info_results[books.tag] = books.text
            odbsci = books.text

        if (books.tag == 'msisdn'):
            msisdn_info_results[books.tag] = books.text
            msisdn = books.text

        if (books.tag == 'isActiveIMSI'):
            msisdn_info_results[books.tag] = books.text
            isActiveIMSI = books.text

        if (books.tag == 'actIMSIGprs'):
            msisdn_info_results[books.tag] = books.text
            actIMSIGprs = books.text

        if (books.tag == 'obGprs'):
            msisdn_info_results[books.tag] = books.text
            obGprs = books.text

        if (books.tag == 'qosProfile'):
            msisdn_info_results[books.tag] = books.text
            qosProfile = books.text

        if (books.tag == 'imeisv'):
            msisdn_info_results[books.tag] = books.text
            imeisv = books.text

        if (books.tag == 'ldapResponse'):
            msisdn_info_results[books.tag] = books.text
            msisdn = books.text

        if (books.tag == 'pdpContext'):
            for attr in books:
                if inc == 1:
                    if (attr.tag == "id"):
                        print(attr.text)
                    if (attr.tag == "refPdpContextName"):
                        msisdn_info_results[attr.tag] = attr.text
                        refPdpContextName = attr.text
                if inc == 2:
                    if (attr.tag == "id"):
                        print(attr.text)
                    if (attr.tag == "refPdpContextName"):
                        print(attr.text)
                if inc == 3:
                    if (attr.tag == "id"):
                        print(attr.text)
                    if (attr.tag == "refPdpContextName"):
                        print(attr.text)
            inc += 1

    # ---------------------------------------------------------------------------------------------------------------
    # ----------------- collectes des donnees dans le fichier excel -------------------------------------------------

    # -----------------------------------------------------------------------------
    # --------Generation du fichier Excel (Avec lignes et colonnes)----------------
    # workbook = openpyxl.Workbook()
    # sheet = workbook.active
    # column = 1
    # for key,values in msisdn_info_results.items():
    #     sheet.cell(row=1, column=column, value=key)
    #     row = 2
    #     sheet.cell(row=row, column=column, value=values)
    #     column += 1
    # workbook.save(filename='Complaints_internet/dataset_internet.xlsx')
    # -----------------------------------------------------------------------------
    # ----------------- Ajout des donnees a la fin du fichier excel----------------
    
    # try:
    filename='Complaints_internet/dataset_internet.xlsx'
    wb = openpyxl.load_workbook(filename=filename)

    sheet = wb.active
    new_row = []

    for key,values in msisdn_info_results.items():
        new_row.append(values)

    sheet.append(new_row)
    wb.save(filename)
    # except :
    #     messageErreur = 'Error -> file not closed:-) You must first closed the "dataset_internet.xlsx" file !'
    #     return messageErreur
    
    # return msisdn_info_results
    # return imsi,encKey,algoId,kdbId,acsub,imsiActive,accTypeGSM,accTypeGERAN,accTypeUTRAN,odboc,odbic,odbr,odboprc,odbssm,odbgprs,odbsci,isActiveIMSI,msisdn,actIMSIGprs,obGprs,qosProfile,refPdpContextName,imeisv,ldapResponse

    return msisdn_info_results
    
if __name__ == "__main__":
    file='soap.xml'
    msisdn_info_results = {}
    put_data_in_dataset(file, msisdn_info_results)