import xml.etree.cElementTree as ET
import openpyxl

def put_data_in_dataset(filename, msisdn_info_results):
    
    """ Return dictionnary of xml value"""

    imsi = None
    encKey = None
    algoId = None
    kdbId = None
    acsub = None
    imsiActive = None
    accTypeGSM = None
    accTypeGERAN = None
    accTypeUTRAN = None
    odboc = None
    odbic = None
    odbr = None
    odboprc = None
    odbssm = None
    odbgprs = None
    odbsci = None
    isActiveIMSI = None
    msisdn = None
    actIMSIGprs = None
    obGprs = None
    qosProfile = None
    refPdpContextName = None
    imeisv = None
    ldapResponse = None

    # Extraction du contenu du fichier xml
    tree = ET.ElementTree(file=filename)
    root = tree.getroot()
    for books in root.findall('.//'):

        for attr in books:
            if (attr.tag == 'imsi'):
                msisdn_info_results[attr.tag] = attr.text
                imsi = attr.text
            
            if (attr.tag == 'encKey'):
                msisdn_info_results[attr.tag] = attr.text
                encKey = attr.text
            
            if (attr.tag == 'algoId'):
                msisdn_info_results[attr.tag] = attr.text
                algoId = attr.text
            
            if (attr.tag == 'kdbId'):
                msisdn_info_results[attr.tag] = attr.text
                kdbId = attr.text
            
            if (attr.tag == 'acsub'):
                msisdn_info_results[attr.tag] = attr.text
                acsub = attr.text
            
            if (attr.tag == 'imsiActive'):
                msisdn_info_results[attr.tag] = attr.text
                imsiActive = attr.text
            
            if (attr.tag == 'accTypeGSM'):
                msisdn_info_results[attr.tag] = attr.text
                accTypeGSM = attr.text
            
            if (attr.tag == 'accTypeGERAN'):
                msisdn_info_results[attr.tag] = attr.text
                accTypeGERAN = attr.text
            
            if (attr.tag == 'accTypeUTRAN'):
                msisdn_info_results[attr.tag] = attr.text
                accTypeUTRAN = attr.text

            if (attr.tag == 'odboc'):
                msisdn_info_results[attr.tag] = attr.text
                odboc = attr.text

            if (attr.tag == 'odbic'):
                msisdn_info_results[attr.tag] = attr.text
                odbic = attr.text

            if (attr.tag == 'odbr'):
                msisdn_info_results[attr.tag] = attr.text
                odbr = attr.text

            if (attr.tag == 'odboprc'):
                msisdn_info_results[attr.tag] = attr.text
                odboprc = attr.text

            if (attr.tag == 'odbssm'):
                msisdn_info_results[attr.tag] = attr.text
                odbssm = attr.text
            
            if (attr.tag == 'odbgprs'):
                msisdn_info_results[attr.tag] = attr.text
                odbgprs = attr.text
            
            if (attr.tag == 'odbsci'):
                msisdn_info_results[attr.tag] = attr.text
                odbsci = attr.text

            if (attr.tag == 'msisdn'):
                msisdn_info_results[attr.tag] = attr.text
                msisdn = attr.text

            if (attr.tag == 'isActiveIMSI'):
                msisdn_info_results[attr.tag] = attr.text
                isActiveIMSI = attr.text
                
            if (attr.tag == 'actIMSIGprs'):
                msisdn_info_results[attr.tag] = attr.text
                actIMSIGprs = attr.text
            
            if (attr.tag == 'obGprs'):
                msisdn_info_results[attr.tag] = attr.text
                obGprs = attr.text
            
            if (attr.tag == 'qosProfile'):
                msisdn_info_results[attr.tag] = attr.text
                qosProfile = attr.text
            
            if (attr.tag == 'refPdpContextName'):
                msisdn_info_results[attr.tag] = attr.text
                refPdpContextName = attr.text
            
            if (attr.tag == 'imeisv'):
                msisdn_info_results[attr.tag] = attr.text
                imeisv = attr.text
                
            if (attr.tag == 'ldapResponse'):
                msisdn_info_results[attr.tag] = attr.text
                msisdn = attr.text
    
    # for value , items in msisdn_info_results.items():
        # print(value + ' = ' + items)
    # ---------------------------------------------------------------------------------------------------------------
    # ----------------- collectes des donnees dans le fichier excel -------------------------------------------------

    # -----------------------------------------------------------------------------
    # --------Generation du fichier Excel (Avec lignes et colonnes)----------------
    # workbook = openpyxl.Workbook()
    # sheet = workbook.active
    # column = 1
    # for key,values in msisdn_info_results.items():
    #     sheet.cell(row=1, column=column, value=key)
    #     row = 2
    #     sheet.cell(row=row, column=column, value=values)
    #     column += 1
    # workbook.save(filename='HLR/dataset_internet.xlsx')
    # -----------------------------------------------------------------------------
    # ----------------- Ajout des donnees a la fin du fichier excel----------------
    
    try:
        filename='HLR/dataset_internet.xlsx'
        wb = openpyxl.load_workbook(filename=filename)

        sheet = wb.active
        new_row = []

        for key,values in msisdn_info_results.items():
            new_row.append(values)

        sheet.append(new_row)
        wb.save(filename)
    except :
        messageErreur = 'Error -> file not closed:-) You must first closed the "dataset_internet.xlsx" file !'
        return messageErreur
    
    # return msisdn_info_results
    # return imsi,encKey,algoId,kdbId,acsub,imsiActive,accTypeGSM,accTypeGERAN,accTypeUTRAN,odboc,odbic,odbr,odboprc,odbssm,odbgprs,odbsci,isActiveIMSI,msisdn,actIMSIGprs,obGprs,qosProfile,refPdpContextName,imeisv,ldapResponse
    
if __name__ == "__main__":
    file='soap.xml'
    msisdn_info_results = {}
    get_in_dataset(file, msisdn_info_results)